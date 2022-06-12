try:
    from openpyxl import load_workbook
except:
    import os
    os.system('pip install openpyxl')
    from openpyxl import load_workbook
import time,os
from urllib import request
import socket
socket.setdefaulttimeout(30)

def mkdir(path):
    folder = os.path.exists(path)
    if not folder:
        os.makedirs(path)
    else:
        print("there is this folder")
        
def save_pic_url(url,path):
    try:
        data = request.urlopen(url,timeout=5).read()
        with open(path,'wb') as f:
            f.write(data)
    except request.URLError as e:
        print('error',e)
    except socket.timeout:
        print('timeouterror')
    except :
        print('unknown error')
        
        
ExcelFullName = './10.xlsx'
wb = load_workbook(ExcelFullName)

for table in wb:
# 循环每一行
    for row in range(1,table.max_row+1):
        dir1 = table.cell(row=row,column=1).value #第一列
        dir2 = table.cell(row=row,column=2).value #第二列
        pic_url = table.cell(row=row,column=3).value #第三列
        #print(dir1)
        mypath=os.path.join(dir1,dir2)
        print(mypath)
        mkdir(mypath)
        #print(pic_url)
        pic_url_list=pic_url.split('\r\n')
        #去除重复的url
        pic_url_list=list(set(pic_url_list))
        for pic in pic_url_list:
            pic_file_name=os.path.basename(pic)
            pic_path=os.path.join(mypath,pic_file_name)
            print(pic_path)
            save_pic_url(pic,pic_path)
