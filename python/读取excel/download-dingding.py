# -*- coding: utf-8 -*-
"""
@author: sunliguo
@contact: QQ376440229
@Created on: 2020/09/08 8:13
1:2020-09-13 添加debug变量
"""
try:
    from openpyxl import load_workbook

except:
    import os

    os.system('pip install openpyxl')
    from openpyxl import load_workbook

import time, os
from urllib import request
import socket

socket.setdefaulttimeout(30)

table_line = 1

debug = 0

# 要处理的execl文件
ExcelFullName = './10.xlsx'


def mkdir(path):
    folder = os.path.exists(path)
    if not folder:
        os.makedirs(path)
    else:
        print("there is this folder")


def save_pic_url(url, path):
    try:
        data = request.urlopen(url, timeout=5).read()
        with open(path, 'wb') as f:
            f.write(data)
    except request.URLError as e:
        print('error', e)
    except socket.timeout:
        print('timeouterror')


wb = load_workbook(ExcelFullName)
for table in wb:
    # 循环每一行

    for row in range(1, table.max_row + 1):
        print("################第{0}行####################".format(table_line))
        dir1 = table.cell(row=row, column=1).value.replace(" ", "").replace("\t", "").strip()  # 第一列
        dir2 = table.cell(row=row, column=2).value.replace(" ", "").replace("\t", "").strip()  # 第二列
        pic_url = table.cell(row=row, column=3).value  # 第三列

        '''
        生成目录
        '''
        mypath = os.path.join(dir1, dir2)
        if not os.path.exists(mypath):
            mkdir(mypath)
        '''
        处理图片url列表
        '''
        # print(pic_url)
        # pic_url = pic_url.replace('\r',"").replace('\n','')
        pic_url_list = pic_url.split('\n')
        # 去除重复的url
        pic_url_list = list(set(pic_url_list))
        while '' in pic_url_list:
            pic_url_list.remove('')
        if debug == 1:
            print('处理后的pic_url_list\n', pic_url_list)
        print("共有{0}张照片".format(len(pic_url_list)))

        # 遍历下载图片
        for pic in pic_url_list:
            pic_file_name = os.path.basename(pic)
            pic_path = os.path.join(mypath, pic_file_name)
            # print(pic_path)
            print("pic url:", pic)
            save_pic_url(pic, pic_path)

            #判断文件是否下载成功
            if os.path.isfile(pic_path):
                print(pic_path, "下载成功")
            else:
                print(pic_path, "下载失败！")

        table_line += 1
